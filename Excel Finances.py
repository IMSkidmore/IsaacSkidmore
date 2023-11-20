import tkinter as tk
from tkinter import simpledialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import subprocess

def on_submit():
    # Get user input for 10 financial categories
    categories = [
        "Food", "Transportation", "Childcare", "Category 4", "Category 5",
        "Category 6", "Category 7", "Category 8", "Category 9", "Category 10"
    ]

    financial_data = {}
    for category in categories:
        value = simpledialog.askfloat("Input", f"Enter value for {category}:")

        if value is not None:  # User can press Cancel to skip a category
            financial_data[category] = value
        else:
            label.config(text="Operation canceled. No data saved.")
            return

    # Update Excel spreadsheet
    file_path = "financial_data.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the current date
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Find the next available column
    next_column = sheet.max_column + 1

    # Label the column with the current date
    label_cell = sheet.cell(row=1, column=next_column)
    label_cell.value = current_date
    label_cell.alignment = Alignment(horizontal='center')  # Center-align the date

    # Save financial data in the labeled column
    for i, category in enumerate(categories, start=2):
        sheet.cell(row=i, column=next_column).value = financial_data.get(category, 0)

    workbook.save(file_path)

    label.config(text=f"Data saved in '{file_path}'. Opening Excel...")

    # Automatically open the Excel file
    subprocess.Popen(["start", "excel", file_path], shell=True)

# Create the main window
app = tk.Tk()
app.title("Financial Data Entry")

# Create a label and a button
label = tk.Label(app, text="Enter financial data for 10 categories.")
label.pack(pady=10)

submit_button = tk.Button(app, text="Submit", command=on_submit)
submit_button.pack(pady=10)

# Run the main loop
app.mainloop()
